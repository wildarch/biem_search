import arff # This should be liac-arff!
from biemsearch import biem_search
from biem.biem import biem
from openpyxl import Workbook

from evaluator.match_ratio import MatchRatioEvaluator
from evaluator.wracc import WraccEvaluator
from evaluator.sensitivity import SensitivityEvaluator
from evaluator.specificity import SpecificityEvaluator
from evaluator.chisquared import ChisquaredEvaluator

# First check that we have the right arff library
if arff.__author__ != 'Renato de Pontes Pereira':
    raise ImportError('You may have imported the wrong arff parsing library, please download liac-arff')


def run_evaluator(evaluator, data, attributes):
    total_matches = 0
    for entry in data:
        if entry[-1] == '1':
            total_matches += 1
    results = biem_search(data, attributes, 10, 2, evaluator)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Name"
    ws['B1'] = evaluator.name()
    ws['C1'] = "Matches"
    ws['D1'] = "Missed matches"
    ws['E1'] = "Size"
    ws['F1'] = "Relative size"

    for (attributes, attr_values, subgroup) in results:
        matches_subgroup = 0
        for index in subgroup:
            if data[index][-1] == '1':
                matches_subgroup += 1
        quality = evaluator.evaluate((attributes, attr_values, subgroup))
        relative_size = len(subgroup)/len(data)

        names = []
        for (name, value) in zip(attributes, attr_values):
            names.append("{0}({1})".format(name, value))
        name = "+".join(names)

        print("{0}: {1} {2:.2f}%, size {3:.2f}% of full data set".format(name, evaluator.name(), quality*100, relative_size*100))
        ws.append([name, quality, matches_subgroup, total_matches - matches_subgroup, len(subgroup), relative_size])

    wb.save("{0}.xlsx".format(evaluator.name()))
    biem()

dataset = arff.load(open('./SpeedDating1-filtered-nocommas-discrete-withouthasnull.arff'))
data = dataset['data']
attributes = dataset['attributes']

match_ratio = MatchRatioEvaluator(data, 0.1)
wracc = WraccEvaluator(data)
sensitivity = SensitivityEvaluator(data, 0.1)
specificity = SpecificityEvaluator(data, 0.1)
chisquared = ChisquaredEvaluator(data, 0.1);

evaluators = [match_ratio, wracc, sensitivity, specificity, chisquared]
for evaluator in evaluators:
    run_evaluator(evaluator, data, attributes)